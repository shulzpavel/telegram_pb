"""
File parser service implementation
"""
import logging
from typing import List
import tempfile
import os

from core.interfaces import IFileParser
from core.exceptions import FileParseError

logger = logging.getLogger(__name__)


class FileParserService(IFileParser):
    """Service for parsing files into tasks"""
    
    def parse_text(self, text: str) -> List[str]:
        """Parse text into tasks"""
        try:
            if not text or not text.strip():
                raise FileParseError("Text is empty")
            
            lines = text.strip().split('\n')
            tasks = []
            
            for line in lines:
                line = line.strip()
                if line and not line.startswith('#'):  # Skip empty lines and comments
                    # Remove markdown list markers
                    if line.startswith(('- ', '* ', '+ ')):
                        line = line[2:]
                    elif line.startswith(('1. ', '2. ', '3. ', '4. ', '5. ', '6. ', '7. ', '8. ', '9. ')):
                        # Remove numbered list markers
                        line = line[3:]
                    
                    if line:
                        tasks.append(line)
            
            if not tasks:
                raise FileParseError("No valid tasks found in text")
            
            logger.info(f"Parsed {len(tasks)} tasks from text")
            return tasks
            
        except Exception as e:
            logger.error(f"Error parsing text: {e}")
            raise FileParseError(f"Failed to parse text: {e}")
    
    def parse_xlsx(self, file_path: str) -> List[str]:
        """Parse xlsx file into tasks"""
        try:
            from openpyxl import load_workbook
            
            if not os.path.exists(file_path):
                raise FileParseError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            if worksheet is None:
                workbook.close()
                raise FileParseError("No active worksheet found")
            
            tasks = []
            headers = []
            
            # Read headers
            for col in range(1, 10):
                cell_value = worksheet.cell(row=1, column=col).value
                headers.append(str(cell_value).lower().strip() if cell_value else "")
            
            # Find key and summary columns
            key_col = None
            summary_col = None
            
            for i, header in enumerate(headers):
                if any(keyword in header for keyword in ['ключ', 'key', 'id', 'номер']):
                    key_col = i + 1
                elif any(keyword in header for keyword in ['резюме', 'summary', 'описание', 'задача', 'название']):
                    summary_col = i + 1
            
            # Parse tasks
            if key_col is None and summary_col is None:
                # Fallback: use columns 2 and 3
                for row_num in range(2, worksheet.max_row + 1):
                    key_cell = worksheet.cell(row=row_num, column=2).value
                    summary_cell = worksheet.cell(row=row_num, column=3).value
                    
                    if key_cell and summary_cell:
                        key_text = str(key_cell).strip()
                        summary_text = str(summary_cell).strip()
                        
                        if (key_text and summary_text and
                            not key_text.lower().startswith(('ключ', 'key', 'id', 'номер')) and
                            not summary_text.lower().startswith(('резюме', 'summary', 'описание', 'задача'))):
                            task_text = f"[{key_text}] {summary_text}"
                            tasks.append(task_text)
            else:
                # Use detected columns
                for row_num in range(2, worksheet.max_row + 1):
                    key_cell = worksheet.cell(row=row_num, column=key_col).value if key_col else None
                    summary_cell = worksheet.cell(row=row_num, column=summary_col).value if summary_col else None
                    
                    if key_cell and summary_cell:
                        key_text = str(key_cell).strip()
                        summary_text = str(summary_cell).strip()
                        
                        if key_text and summary_text:
                            task_text = f"[{key_text}] {summary_text}"
                            tasks.append(task_text)
            
            workbook.close()
            
            if not tasks:
                raise FileParseError("No valid tasks found in xlsx file")
            
            logger.info(f"Parsed {len(tasks)} tasks from xlsx file")
            return tasks
            
        except ImportError:
            raise FileParseError("openpyxl library not installed")
        except Exception as e:
            logger.error(f"Error parsing xlsx file: {e}")
            raise FileParseError(f"Failed to parse xlsx file: {e}")
    
    def parse_file(self, file_path: str) -> List[str]:
        """Parse file based on extension"""
        try:
            if file_path.lower().endswith(('.xlsx', '.xls')):
                return self.parse_xlsx(file_path)
            else:
                # Try to read as text
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                return self.parse_text(content)
                
        except Exception as e:
            logger.error(f"Error parsing file: {e}")
            raise FileParseError(f"Failed to parse file: {e}")
